import pathlib

import pandas as pd
import sqlite3

from openpyxl.styles import Font
from pyecharts.charts import Bar, Line, Tab, Grid, Page, Pie
import pyecharts.options as opts
from pyecharts.components import Table
from snapshot_phantomjs import snapshot
from pyecharts.globals import ThemeType
from pyecharts.render import make_snapshot
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl import Workbook
import settings


class HouseDataAnalyse:
    def __init__(self, global_df: pd.DataFrame, target_df: pd.DataFrame, data_date):
        """
        数据分析与可视化
        :param global_df: 全量清洗数据
        :param target_df: 目标数据
        :param data_date: 数据日期
        """
        self.global_df = global_df
        self.target_df = target_df
        self.data_date = data_date
        self.db_path = settings.db_path
        self.picture_path = settings.picture_path
        self.report_path = settings.report_path

    @staticmethod
    def data_store(table_name):
        pass

    @staticmethod
    def data_analyse(df: pd.DataFrame):
        # 区域名称
        region_name_list = []
        # 区域二手房总数
        region_housing_num_list = []
        # 区域房价最高值
        region_housing_max_list = []
        # 区域房价最低值
        region_housing_min_list = []
        # 区域房价平均值
        region_housing_avg_list = []
        # 区域房价中位数
        region_housing_median_list = []
        # 区域房价标准差
        region_housing_std_list = []
        # 区域房价方差
        region_housing_var_list = []

        city_mean_price = round(df['housing_unit_price'].mean(), 2)
        all_list = []
        groups = df.groupby("city_region")
        for group_name, group_df in groups:
            f_se = group_df['housing_unit_price'].agg(['max', 'min', 'mean', 'median', 'std', 'var'])
            f_count = group_df['city_region'].agg("count")
            all_list.append([city_mean_price, group_name, f_count, f_se[0], f_se[1], round(f_se[2], 2), f_se[3]])
            region_name_list.append(group_name)
            region_housing_num_list.append(f_count)
            region_housing_max_list.append(f_se[0])
            region_housing_min_list.append(f_se[1])
            region_housing_avg_list.append(round(f_se[2], 2))
            region_housing_median_list.append(f_se[3])
            region_housing_std_list.append(f_se[4])
            region_housing_var_list.append(f_se[5])

        return all_list, region_name_list, region_housing_num_list, region_housing_max_list, region_housing_min_list, \
               region_housing_avg_list, region_housing_median_list, region_housing_std_list, region_housing_var_list

    def global_data_analyse(self):
        """
        全量数据分析
        1、计算城市平均房价
        2、计算每个区的房价平均值，方差，平均差
        3、计算每个区在售房源
        :return:
        """

        return self.data_analyse(self.global_df)

    def target_df_analyse(self):
        """
        目标数据分析
        1、与昨日相比新增数据
        2、与昨日对比价格不变房源。价格上涨房源，价格下降房源
        2、每个区的每一个符合条件房源价格变化
        :return:
        """
        # 读取sqlite数据库中昨日数据
        # read_sql = "select * from {table_name}".format(table_name=self.yesterday_clean_table_name)
        # with sqlite3.connect(self.db_path) as conn:
        #     c = conn.cursor()
        #     df = pd.read_sql(read_sql, con=c)
        """
        目标数据分析
        1、计算城市平均房价
        2、计算每个区的房价平均值，方差，平均差
        3、计算每个区在售房源
        :return:
        """
        return self.data_analyse(self.target_df)

    def global_df_visual(self):
        """
        全局数据可视化，生成图片
        :return:图片地址
        """
        all_list, region_name_list, region_housing_num_list, region_housing_max_list, region_housing_min_list, \
        region_housing_avg_list, region_housing_median_list, region_housing_std_list, \
        region_housing_var_list = self.global_data_analyse()
        # 各区域二手房数量
        housing_num_bar = (
            Bar(init_opts=opts.InitOpts(theme=ThemeType.ROMA, )).add_xaxis(region_name_list)
                .add_yaxis("", [str(x) for x in region_housing_num_list]).set_global_opts(
                title_opts=opts.TitleOpts(title="各区域二手房数量"),
                yaxis_opts=opts.AxisOpts(name='数量（个）')
            )
        )
        # 单位面积价格统计
        housing_agg_bar = Line(init_opts=opts.InitOpts(theme=ThemeType.ROMA)).add_xaxis(region_name_list).add_yaxis(
            series_name="单位面积价格最大值",
            y_axis=region_housing_max_list
        ).add_yaxis(
            series_name="单位面积价格最小值",
            y_axis=region_housing_min_list
        ).add_yaxis(
            series_name="单位面积价格平均数",
            y_axis=region_housing_avg_list
        ).set_global_opts(
            title_opts=opts.TitleOpts(title="单位面积价格统计"),
            yaxis_opts=opts.AxisOpts(name='价格（元）')
        )
        # 明细数据表格
        # table = Table()
        # headers = ["南京二手房平均价格", "区域名称", "二手房数量", "单位面积价格最大值", "单位面积价格最小值", "单位面积价格平均数", "单位面积价格中位数"]
        #
        # table.add(headers, all_list).set_global_opts(
        #     title_opts=opts.ComponentTitleOpts(title="明细数据表格")
        # )
        # page = Page(layout=Page.SimplePageLayout)
        # page.add(housing_num_bar, housing_agg_bar, table)
        # page.render()
        make_snapshot(snapshot, housing_num_bar.render(), self.picture_path + "{data_date}_housing_num_bar.png"
                      .format(data_date=self.data_date))
        make_snapshot(snapshot, housing_agg_bar.render(), self.picture_path + "{data_date}_housing_agg_bar.png"
                      .format(data_date=self.data_date))
        return (self.picture_path + "{data_date}_housing_num_bar.png".format(
            data_date=self.data_date), self.picture_path + "{data_date}_housing_agg_bar.png".format(
            data_date=self.data_date), all_list)

    def target_df_visual(self):
        """
        目标数据可视化
        :return:
        """
        all_list, region_name_list, region_housing_num_list, region_housing_max_list, region_housing_min_list, \
        region_housing_avg_list, region_housing_median_list, region_housing_std_list, \
        region_housing_var_list = self.target_df_analyse()
        # 各区域二手房数量
        housing_num_bar = (
            Bar(init_opts=opts.InitOpts(theme=ThemeType.ROMA, )).add_xaxis(region_name_list)
                .add_yaxis("", [str(x) for x in region_housing_num_list]).set_global_opts(
                title_opts=opts.TitleOpts(title="各区域二手房数量"),
                yaxis_opts=opts.AxisOpts(name='数量（个）')
            )
        )
        # 单位面积价格统计
        housing_agg_bar = Line(init_opts=opts.InitOpts(theme=ThemeType.ROMA)).add_xaxis(region_name_list).add_yaxis(
            series_name="单位面积价格最大值",
            y_axis=region_housing_max_list
        ).add_yaxis(
            series_name="单位面积价格最小值",
            y_axis=region_housing_min_list
        ).add_yaxis(
            series_name="单位面积价格平均数",
            y_axis=region_housing_avg_list
        ).set_global_opts(
            title_opts=opts.TitleOpts(title="单位面积价格统计"),
            yaxis_opts=opts.AxisOpts(name='价格（元）')
        )
        make_snapshot(snapshot, housing_num_bar.render(), self.picture_path + "{data_date}_target_housing_num_bar.png"
                      .format(data_date=self.data_date))
        make_snapshot(snapshot, housing_agg_bar.render(), self.picture_path + "{data_date}_target_housing_agg_bar.png"
                      .format(data_date=self.data_date))
        city_mean_price = all_list[0][0]
        city_housing_num = sum(region_housing_num_list)
        return (self.picture_path + "{data_date}_target_housing_num_bar.png".format(
            data_date=self.data_date), self.picture_path + "{data_date}_target_housing_agg_bar.png".format(
            data_date=self.data_date), city_mean_price, city_housing_num)

    def create_report(self):
        """
        结合dataframe和图片生成excel报告文档
        :return:
        """
        title_font = Font(name=u'微软雅黑', bold=True, italic=False, size=24)
        text_font = Font(name=u'微软雅黑', bold=False, italic=False, size=20)
        imgsize = (1280 / 1.5, 720 / 1.5)
        housing_num_bar_path, housing_agg_bar_path, all_list = self.global_df_visual()
        housing_num_bar_img = Image(housing_num_bar_path)
        housing_agg_bar_img = Image(housing_agg_bar_path)
        housing_num_bar_img.width, housing_num_bar_img.height = imgsize
        housing_agg_bar_img.width, housing_agg_bar_img.height = imgsize
        report_excel_path = self.report_path + "{data_date}二手房行情.xlsx".format(data_date=self.data_date)
        if not pathlib.Path(report_excel_path).exists():
            report_excel_wb = Workbook()
        else:
            report_excel_wb = load_workbook(report_excel_path)
        if 'Sheet' in report_excel_wb.sheetnames:
            report_excel_wb['Sheet'].title = "南京整体房价分析"
        # ws = report_excel_wb.create_sheet("南京整体房价分析", 0)
        # ws.sheet_properties.tabColor = 'ff72BA'

        ws = report_excel_wb["南京整体房价分析"]
        ws['A1'].font = title_font
        ws['A1'] = '南京整体房价分析'
        ws['A2'].font = text_font
        ws['A2'] = "平均房价为{a}".format(a=all_list[0][0])
        ws.column_dimensions['A'].width = imgsize[0] * 0.14  # 修改列A的宽
        ws.column_dimensions['B'].width = imgsize[0] * 0.14  # 修改列B的宽
        ws.add_image(housing_num_bar_img, 'A3')  # 添加图片到指定的单元格
        ws.row_dimensions[3].height = imgsize[1] * 0.78
        ws.add_image(housing_agg_bar_img, 'B3')  # 添加图片到指定的单元格
        ws.row_dimensions[3].height = imgsize[1] * 0.78
        target_ws = report_excel_wb.create_sheet("南京符合条件二手房分析", 1)

        target_ws['A1'].font = title_font
        target_ws['A2'].font, target_ws['B2'].font = text_font, text_font
        target_ws['A1'] = '南京符合要求房价分析'
        target_housing_num_bar_path, target_housing_agg_bar_path, city_mean_price, city_housing_num = \
            self.target_df_visual()
        target_housing_num_bar_img = Image(target_housing_num_bar_path)
        target_housing_agg_bar_img = Image(target_housing_agg_bar_path)
        target_housing_num_bar_img.width, target_housing_num_bar_img.height = imgsize
        target_housing_agg_bar_img.width, target_housing_agg_bar_img.height = imgsize
        target_ws['A2'] = "平均房价为{a}".format(a=city_mean_price)
        target_ws['B2'] = "符合条件二手房数量为{a}".format(a=city_housing_num)
        target_ws.column_dimensions['A'].width = imgsize[0] * 0.14  # 修改列A的宽
        target_ws.column_dimensions['B'].width = imgsize[0] * 0.14  # 修改列B的宽
        target_ws.add_image(target_housing_num_bar_img, 'A3')  # 添加图片到指定的单元格
        target_ws.row_dimensions[3].height = imgsize[1] * 0.78
        target_ws.add_image(target_housing_agg_bar_img, 'B3')  # 添加图片到指定的单元格
        target_ws.row_dimensions[3].height = imgsize[1] * 0.78

        report_excel_wb.save(report_excel_path)
        report_excel_wb.close()
        book = load_workbook(report_excel_path)
        writer = pd.ExcelWriter(report_excel_path, engine='openpyxl')
        writer.book = book
        target_df = self.target_df
        target_df.rename(columns=settings.column_dict, inplace=True)
        target_df.to_excel(excel_writer=writer, sheet_name='南京符合要求数据', index=False)
        writer.save()
        writer.close()


if __name__ == '__main__':
    df01 = pd.read_csv("../csv_store/2022-05-01处理过链家数据.csv")
    df02 = pd.read_csv("../csv_store/2022-05-01目标链家数据.csv")
    houseDataAnalyse = HouseDataAnalyse(global_df=df01, target_df=df02, data_date='2022-05-03')
    houseDataAnalyse.create_report()
